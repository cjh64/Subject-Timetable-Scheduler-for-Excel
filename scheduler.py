import random
import pandas as pd
import sys
import os
from deap import base, creator, tools, algorithms
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Ensure script runs in the same directory as the Excel file
script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))  # Get the directory of the executable
os.chdir(script_dir)

# Load Excel Data
excel_path = os.path.join(script_dir, "test.xlsm")  # Absolute path to the Excel file
wb = load_workbook(excel_path, data_only=True)

# Load Subject Sheet
ws_subject = wb["Subject"]
df_subjects = pd.DataFrame(ws_subject.values)
df_subjects.columns = df_subjects.iloc[0].str.strip()  # Ensure column names are stripped
df_subjects = df_subjects[1:].reset_index(drop=True)   # Remove header row from data

# Now create SubjectID -> SubjectName mapping
subject_name_dict = df_subjects.set_index("SubjectID")["SubjectName"].to_dict()

# Assuming 'NoStudent' column might have NaN values
df_subjects["NoStudent"] = pd.to_numeric(df_subjects["NoStudent"], errors="coerce").fillna(0).astype(int)

# Load Venue and Lecturer Sheets
df_venues = pd.read_excel(excel_path, sheet_name="Venue")
df_lecturers = pd.read_excel(excel_path, sheet_name="Lecturer", header=None)

# Process Lecturer Sheet
for index, row in df_lecturers.iterrows():
    if "LecturerID" in row.values:
        df_lecturers.columns = df_lecturers.iloc[index]
        df_lecturers = df_lecturers[index+1:].reset_index(drop=True)
        break
df_lecturers.columns = df_lecturers.columns.str.strip()
df_lecturers = df_lecturers.dropna(how="all").reset_index(drop=True)

# Extract Data
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
times = list(range(11))  # Time slots from 0 to 10

venue_dict = {
    "Lecture": df_venues[df_venues["Type"] == "Lecture Hall"]["Venue"].tolist(),
    "Tutorial": df_venues[df_venues["Type"] == "Tutorial Room"]["Venue"].tolist(),
    "Lab": df_venues[df_venues["Type"] == "Lab"]["Venue"].tolist(),
}

venue_capacities = df_venues.set_index("Venue")["Capacity"].to_dict()
venue_types = df_venues.set_index("Venue")["Type"].to_dict()
lecturer_dict = df_lecturers.set_index("LecturerID")["LecturerName"].to_dict()
lecturers = list(lecturer_dict.keys())

# Store venue and lecturer availability
venue_schedule = {venue: [] for venue in sum(venue_dict.values(), [])}  # Initialize all venues as empty
lecturer_subject_count = {lec: 0 for lec in lecturers}  

def get_time_range(start_slot, duration):
    start_hour = 8 + start_slot  
    end_hour = start_hour + duration
    return f"{start_hour}:00", f"{end_hour}:00"

# Venue Selection Function with Venue Utilization and Clash Check (End time constraint added)
def select_venue(session_type, student_count, session_time, session_day):
    """ Selects an available venue that fits the session type and capacity constraints, checking for time clashes. """
    available_venues = [v for v in venue_dict[session_type] if venue_capacities.get(v, 0) >= student_count]
    
    if not available_venues:
        return "No Venue"
    
    # Sort the venues by usage count (less used venues are prioritized)
    sorted_venues = sorted(available_venues, key=lambda v: len(venue_schedule[v]))

    for venue in sorted_venues:
        # Check for time slot clashes in the selected venue
        clash_found = False
        for scheduled_session in venue_schedule[venue]:
            existing_day, existing_start, existing_end = scheduled_session
            # Ensure the new session does not start before the existing session ends
            if existing_day == session_day and not (session_time + 2 <= existing_start or session_time >= existing_end):
                clash_found = True
                break
        
        if not clash_found:
            # No clash found, so assign the venue
            venue_schedule[venue].append((session_day, session_time, session_time + 2))
            return venue  # Venue found with no clash
    
    return "No Venue"  # Return if no valid venue found


def split_students(subject_id, major, total_students, lecture_hours, tutorial_hours, lab_hours):
    sessions = []
    
    major = str(major)
    # Get valid lecturers based on the major and ensure they teach max 3 subjects
    valid_lecturers = [
        lec for lec in df_lecturers[df_lecturers["Major"].str.contains(major, na=False)]["LecturerID"].tolist()
        if lecturer_subject_count[lec] < 3
    ]

    if len(valid_lecturers) < 2:
        return []  

    num_lecturers = min(3, len(valid_lecturers))
    subject_lecturers = random.sample(valid_lecturers, num_lecturers)  # Select lecturers
    main_lecturer = random.choice(subject_lecturers)  # Choose the main lecturer for the subject

    # Assign a lecturer to the subject
    for lec in subject_lecturers:
        lecturer_subject_count[lec] += 1  

    # Assign Lecture Sessions (Split if 3 hours)
    if lecture_hours > 0:
        session_time = random.choice(times)  # Ensure session_time is a valid integer

        lecture_venue = select_venue("Lecture", total_students, session_time, None)  # Assign lecture venue here

        if total_students > 300:
            first_half = total_students // 2
            second_half = total_students - first_half

            sessions.append((subject_id, "Lecture", first_half, 2, main_lecturer, lecture_venue))
            sessions.append((subject_id, "Lecture", second_half, 2, main_lecturer, lecture_venue))
        else:
            if lecture_hours == 3:
                sessions.append((subject_id, "Lecture", total_students, 2, main_lecturer, lecture_venue))
                sessions.append((subject_id, "Lecture", total_students, 1, main_lecturer, lecture_venue))
            else:
                sessions.append((subject_id, "Lecture", total_students, lecture_hours, "", lecture_venue))

    # Create Tutorial/Lab Sessions with different random venues for each session
    def create_sessions(session_type, total_students, hours):
        groups = (total_students // 40) + (1 if total_students % 40 != 0 else 0)
        base_size = total_students // groups
        remainder = total_students % groups
        group_sizes = [base_size + (1 if i < remainder else 0) for i in range(groups)]

        for size in group_sizes:
            session_time = random.choice(times)  # Ensure session_time is a valid integer
            session_venue = select_venue(session_type, size, session_time, None)  # Different venue for tutorial/lab
            session_lecturer = random.choice(subject_lecturers)  # Randomly assign a lecturer
            sessions.append((subject_id, session_type, size, hours, "", session_venue))  # Assign session

    # Tutorial and Lab sessions
    if tutorial_hours > 0:
        create_sessions("Tutorial", total_students, tutorial_hours)
    if lab_hours > 0:
        create_sessions("Lab", total_students, lab_hours)

    return sessions

expanded_subjects = []
for _, row in df_subjects.iterrows():
    expanded_subjects.extend(
        split_students(row["SubjectID"], row["Major"], int(row["NoStudent"]), 
                       int(row["Lecture"]) if pd.notna(row["Lecture"]) else 0, 
                       int(row["Tutorial"]) if pd.notna(row["Tutorial"]) else 0, 
                       int(row["Lab"]) if pd.notna(row["Lab"]) else 0)
    )

# ---------------------- Genetic Algorithm Setup ----------------------
creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
creator.create("Individual", list, fitness=creator.FitnessMin)
toolbox = base.Toolbox()

def create_individual():
    individual = []
    venue_schedule = {}  # Track venue usage as {venue: [(day, start_time, end_time)]}

    for subject_id, session_type, student_count, hours, lecturer, _ in expanded_subjects:
        while True:
            day = random.choice(days)
            
            # Define allowed end times based on day
            if day != "Friday":
                # For non-Friday days: classes must end at 10,12,14,16,18
                allowed_end_times = [10, 12, 14, 16, 18]
            else:
                # For Friday: classes must end at 10,12,17,19
                allowed_end_times = [10, 12, 17, 19]
            
            # Find possible time slots that would result in these end times
            possible_slots = []
            for end_time in allowed_end_times:
                # Calculate possible start times that would end at this time
                start_time = end_time - hours
                if start_time >= 8:  # Earliest class starts at 8am
                    time_slot = start_time - 8  # Convert to time slot index (0-10)
                    if 0 <= time_slot <= 10:  # Ensure it's a valid time slot
                        possible_slots.append(time_slot)
            
            if not possible_slots:
                continue  # Skip if no valid slots for this session length
            
            time = random.choice(possible_slots)
            session_start = 8 + time
            session_end = session_start + hours

            # Select a venue that fits capacity and doesn't have time clashes
            available_venues = []
            for venue in venue_dict[session_type]:
                if venue_capacities.get(venue, 0) >= student_count:
                    # Check for time clashes in this venue
                    clash_found = False
                    if venue in venue_schedule:
                        for scheduled_day, scheduled_start, scheduled_end in venue_schedule[venue]:
                            if day == scheduled_day:
                                if not (session_end <= scheduled_start or session_start >= scheduled_end):
                                    clash_found = True
                                    break
                    if not clash_found:
                        available_venues.append(venue)

            if available_venues:
                venue = random.choice(available_venues)
                # Track this venue usage
                if venue not in venue_schedule:
                    venue_schedule[venue] = []
                venue_schedule[venue].append((day, session_start, session_end))
                break  # Valid slot found

        individual.append((day, time, venue, "", hours))
    return creator.Individual(individual)
    
toolbox.register("individual", create_individual)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)

def evaluate(individual):
    penalty = 0
    lecturer_schedule = {}
    venue_schedule = {}  # This will track venue usage as {venue: [(day, start_time, end_time)]}
    subject_venue_map = {}

    for i, entry in enumerate(individual):
        day, time, venue, lecturer, hours = entry
        subject_id, session_type, student_count, _, _, _ = expanded_subjects[i]
        session_start = 8 + time
        session_end = session_start + hours

        # Hard Constraint: Check allowed end times
        if day != "Friday":
            if session_end not in [10, 12, 14, 16, 18]:
                penalty += 1000  # Heavy penalty for violating end time constraint
        else:
            if session_end not in [10, 12, 17, 19]:
                penalty += 1000  # Heavy penalty for violating Friday end time constraint

        # Rest of the evaluation remains the same...
        # Hard Constraint: No classes on Friday 12 PM - 2 PM
        if day == "Friday" and (12 <= session_start < 14):
            penalty += 1000  

        # Hard Constraint: Ensure lectures use the same venue
        if session_type == "Lecture":
            if subject_id in subject_venue_map and subject_venue_map[subject_id] != venue:
                penalty += 50
            else:
                subject_venue_map[subject_id] = venue

        # Venue capacity check
        if venue_capacities.get(venue, 0) < student_count:
            penalty += 50  

        # Avoid lecturer clashes
        lecturer_key = (day, time)
        if lecturer_key in lecturer_schedule:
            penalty += 30  
        else:
            lecturer_schedule[lecturer_key] = lecturer

        # Check for venue time clashes
        if venue in venue_schedule:
            for scheduled_day, scheduled_start, scheduled_end in venue_schedule[venue]:
                if day == scheduled_day:
                    # Check if new session starts before existing session ends
                    if not (session_end <= scheduled_start or session_start >= scheduled_end):
                        penalty += 50  # Add penalty for time clash
                        break
        
        # Track this venue usage
        if venue not in venue_schedule:
            venue_schedule[venue] = []
        venue_schedule[venue].append((day, session_start, session_end))

    return (penalty,)

toolbox.register("mate", tools.cxTwoPoint)
toolbox.register("mutate", tools.mutShuffleIndexes, indpb=0.2)
toolbox.register("select", tools.selTournament, tournsize=3)
toolbox.register("evaluate", evaluate)

def run_ga():
    population = toolbox.population(n=100)
    algorithms.eaSimple(population, toolbox, cxpb=0.5, mutpb=0.2, ngen=50, verbose=True)
    return tools.selBest(population, k=1)[0]

optimized_schedule = run_ga()

output_data = []
for i, entry in enumerate(optimized_schedule):
    day, time_slot, venue, lecturer_id, hours = entry
    subject_id, session_type, student_count, _, _, _ = expanded_subjects[i]
    start_time, end_time = get_time_range(time_slot, hours)
    
    output_data.append((subject_id, subject_name_dict.get(subject_id, "Unknown"), session_type, day, start_time, end_time, venue, student_count, hours))

# Add an empty "Lecturer Name" column at the end of the DataFrame
output_df = pd.DataFrame(output_data, columns=["SubjectID", "SubjectName", "SessionType", "Day", "StartTime", "EndTime", "Venue", "StudentCount", "Hours"])

# Add an empty column for "Lecturer Name"
output_df["LecturerName"] = ""


# ---------------------- Auto-Increment File Saving ----------------------
def get_next_filename(base_name="optimized_schedule", ext=".xlsx"):
    count = 1
    while os.path.exists(f"{base_name}_{count}{ext}"):
        count += 1
    return f"{base_name}_{count}{ext}"

# Save the output DataFrame to a new file with auto-incremented name
output_file = get_next_filename()
output_df.to_excel(output_file, index=False)

# Apply borders and autofit to the new sheet
wb = load_workbook(output_file)
ws = wb.active  # Select the first sheet

# --- Add Excel Table Format ---
from openpyxl.worksheet.table import Table, TableStyleInfo

# Define table range (from A1 to the last cell with data)
min_col = 1
max_col = ws.max_column
min_row = 1
max_row = ws.max_row

table_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
table = Table(displayName="Timetable", ref=table_range)

# Apply a table style
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

# Define border style
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)

# Apply border to all cells in the worksheet
for row in ws.iter_rows():
    for cell in row:
        cell.border = border

# Autofit columns
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # Adjust the width a bit more
    ws.column_dimensions[column].width = adjusted_width

# Save the modified file
wb.save(output_file)

# ---------------------- Visualization and Dashboard ----------------------
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Create a directory for temporary image files
os.makedirs('temp_plots', exist_ok=True)

# Function to save plots as images
def save_plot(fig, filename, dpi=100):
    path = os.path.join('temp_plots', filename)
    fig.savefig(path, bbox_inches='tight', dpi=dpi)
    plt.close(fig)
    return path

# 1. Venue Usage by Type (Top 10)
venue_counts = output_df.groupby(['SessionType', 'Venue']).size().unstack(fill_value=0)
fig1, ax1 = plt.subplots(figsize=(8, 5))
venue_counts.sum().sort_values(ascending=False).head(10).plot(
    kind='bar', ax=ax1, title='Top 10 Most Used Venues'
)
ax1.set_ylabel('Number of Sessions')
venue_usage_path = save_plot(fig1, 'venue_usage.png', dpi=120)

# 2. Session Type Distribution
fig2, ax2 = plt.subplots(figsize=(6, 5))
output_df['SessionType'].value_counts().plot(
    kind='pie', autopct='%1.1f%%', ax=ax2, title='Session Type Distribution'
)
session_dist_path = save_plot(fig2, 'session_distribution.png', dpi=120)

# 3. Daily Session Distribution
fig3, ax3 = plt.subplots(figsize=(8, 5))
output_df.groupby(['Day', 'SessionType']).size().unstack().plot(
    kind='bar', stacked=True, ax=ax3, title='Sessions per Day by Type'
)
ax3.set_ylabel('Number of Sessions')
daily_sessions_path = save_plot(fig3, 'daily_sessions.png', dpi=120)

# 4. Hourly Session Distribution
try:
    output_df['Hour'] = pd.to_datetime(output_df['StartTime'], format='%H:%M').dt.hour
except ValueError:
    output_df['Hour'] = pd.to_datetime(output_df['StartTime']).dt.hour

fig4, ax4 = plt.subplots(figsize=(8, 5))
output_df.groupby(['Hour', 'SessionType']).size().unstack().plot(
    kind='bar', stacked=True, ax=ax4, title='Sessions per Hour by Type'
)
ax4.set_ylabel('Number of Sessions')
ax4.set_xlabel('Starting Hour')
hourly_sessions_path = save_plot(fig4, 'hourly_sessions.png', dpi=120)

#5. Prepare hourly venue usage matrix
heatmap_df = output_df.copy()
heatmap_df['Hour'] = pd.to_datetime(heatmap_df['StartTime'], format='%H:%M').dt.hour
heatmap_matrix = heatmap_df.pivot_table(index='Hour', columns='Day', values='Venue', aggfunc='count', fill_value=0)

# Create heatmap
import seaborn as sns
fig5, ax5 = plt.subplots(figsize=(8, 6))
sns.heatmap(heatmap_matrix, annot=True, fmt="d", cmap="YlGnBu", ax=ax5)
ax5.set_title("Heatmap of Venue Usage by Hour and Day")
heatmap_path = save_plot(fig5, 'venue_heatmap.png', dpi=120)

# 5. Subject-wise Session Count
fig6, ax6 = plt.subplots(figsize=(10, 6))
output_df['SubjectName'].value_counts().sort_values(ascending=True).plot(
    kind='barh', ax=ax6, color='skyblue', title='Sessions per Subject'
)
ax6.set_xlabel('Number of Sessions')
subject_session_path = save_plot(fig6, 'subject_sessions.png', dpi=120)


# Create dashboard sheet with table layout
wb = load_workbook(output_file)
if 'Dashboard' in wb.sheetnames:
    wb.remove(wb['Dashboard'])
wb.create_sheet('Dashboard')
dashboard = wb['Dashboard']

# Set column widths
for col in ['A', 'C']:
    dashboard.column_dimensions[col].width = 55

# Add title
dashboard.merge_cells('A1:C1')
dashboard['A1'] = 'SCHEDULING DASHBOARD'
dashboard['A1'].font = 'Arial Black'
dashboard['A1'].font = '18'
dashboard['A1'].alignment = Alignment(horizontal='center')

# Add images in a 2x2 grid layout
img1 = Image(venue_usage_path)
img1.width, img1.height = 395, 300
dashboard.add_image(img1, 'A3')

img2 = Image(session_dist_path)
img2.width, img2.height = 395, 300
dashboard.add_image(img2, 'C3')

img3 = Image(daily_sessions_path)
img3.width, img3.height = 395, 300
dashboard.add_image(img3, 'A20')

img4 = Image(hourly_sessions_path)
img4.width, img4.height = 395, 300
dashboard.add_image(img4, 'C20')

img5 = Image(heatmap_path)
img5.width, img5.height = 395, 300
dashboard.add_image(img5, 'A37')

img6 = Image(subject_session_path)
img6.width, img6.height = 395, 300
dashboard.add_image(img6, 'C37')

# Add chart titles as headers
dashboard['A2'] = 'Venue Utilization'
dashboard['C2'] = 'Session Distribution'
dashboard['A19'] = 'Daily Schedule Pattern'
dashboard['C19'] = 'Hourly Schedule Pattern'
dashboard['A36'] = 'Venue Usage Heatmap'
dashboard['C36'] = 'Subject-wise Session Count'

# Format headers
for cell in ['A2', 'C2', 'A19', 'C19']:
    dashboard[cell].font = 'Calibri'
    dashboard[cell].font = '14'
    dashboard[cell].font = 'Bold'
    dashboard[cell].alignment = Alignment(horizontal='center')

# Add borders to create table effect
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in range(2, 35):
    for col in ['A', 'C']:
        cell = f"{col}{row}"
        dashboard[cell].border = thin_border

# Save the workbook with dashboard
wb.save(output_file)

# Open the optimized schedule file (Windows)
os.startfile(output_file)

# Clean up temporary plot files
for file in os.listdir('temp_plots'):
    os.remove(os.path.join('temp_plots', file))
os.rmdir('temp_plots')
